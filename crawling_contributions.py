import re
import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, colors


def write_excel_template(filename, idx, sheetname, listdata):
    header_font = Font(bold=True, size=14)
    hyperlink_font = Font(color=colors.BLUE, underline="single")

    try:
        excel_file = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        excel_file = openpyxl.Workbook()
        excel_file.remove(excel_file["Sheet"])

    excel_file.create_sheet(sheetname, idx)
    excel_sheet = excel_file[sheetname]
    excel_sheet.column_dimensions["A"].width = 20
    excel_sheet.column_dimensions["B"].width = 10
    excel_sheet.column_dimensions["C"].width = 10
    excel_sheet.column_dimensions["D"].width = 10
    excel_sheet.column_dimensions["E"].width = 10
    excel_sheet.column_dimensions["F"].width = 70
    excel_sheet.column_dimensions["G"].width = 20
    excel_sheet.column_dimensions["H"].width = 20
    excel_sheet.column_dimensions["I"].width = 30
    excel_sheet.column_dimensions["J"].width = 100

    row = 1
    for item in listdata:
        excel_sheet.append(item)
        if row == 1:
            for cell in excel_sheet[row]:
                cell.font = header_font
        else:
            excel_sheet.cell(row=row, column=10).hyperlink = item[-1]
            excel_sheet.cell(row=row, column=10).font = hyperlink_font
        row += 1
    excel_file.save(filename)
    excel_file.close()


spec_list = ["0uhr", "00bn", "0eht", "00be"]
spec_list2 = [5, 9, 2, 111]
url = "https://mentor.ieee.org/802.11/documents?"

for i in range(4):
    contributions_lists = []
    print("Start crawling contributions: " + spec_list[i])
    for page_num in range(spec_list2[i]):
        print(spec_list[i] + ": " + str(page_num + 1) + "/" + str(spec_list2[i]))
        res = requests.get(url + "n=" + str(page_num + 1) + "&is_group=" + spec_list[i])
        soup = BeautifulSoup(res.content, "html.parser")

        table = soup.select_one("table.paged_list")
        contents = table.find_all("tr")
        if page_num == 0:
            head = contents[0]
            heading = []
            for item in head.find_all("th"):
                item = item.text.strip()
                if item == "Author (Affiliation)":
                    heading.append("Author")
                    heading.append("Affiliation")
                else:
                    heading.append(item)
            heading[-1] = "Download Link"
            contributions_lists.append(heading)

        for tr in contents[1:]:
            element = tr.find_all("td")
            row = []
            tr2_idx = 0
            for tr2 in element:
                tr3 = tr2.text.strip()
                # EHT 1908 LG....^^
                tr3 = re.sub(r"", " ", tr3)

                if tr2_idx == 6:
                    if tr3.find("(") == -1:
                        tr3_1 = ""
                        tr3_2 = tr3
                    else:
                        tr3_1 = tr3.split("(")[0].strip()
                        tr3_2 = tr3.split("(")[1].strip()[:-1]
                    row.append(tr3_1)
                    row.append(tr3_2)
                else:
                    row.append(tr3)
                tr2_idx += 1
            link = "https://mentor.ieee.org" + element[-1].find_all("a")[0].get("href")
            row[-1] = link
            contributions_lists.append(row)
    write_excel_template(
        "802.11_contributions.xlsx", i, spec_list[i], contributions_lists
    )
